from bs4 import BeautifulSoup as bs
from pathlib import Path
from openpyxl import Workbook
from fake_useragent import UserAgent
from requests.exceptions import RequestException, Timeout, ConnectTimeout, ReadTimeout
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.proxy import Proxy, ProxyType
import time
import os
import re
import requests as rq
import math
import sys
import random
import itertools
import json
from urllib.parse import urlencode


class MobileUserAgent:
    """Android, Macintosh, iPhone만 사용하는 User-Agent 생성기"""

    def __init__(self):
        self.user_agents = [
            # Android Chrome
            "Mozilla/5.0 (Linux; Android 13; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
            "Mozilla/5.0 (Linux; Android 12; SM-A525F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Mobile Safari/537.36",
            "Mozilla/5.0 (Linux; Android 11; SM-G973F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Mobile Safari/537.36",
            "Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
            "Mozilla/5.0 (Linux; Android 12; Pixel 6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Mobile Safari/537.36",

            # iPhone Safari
            "Mozilla/5.0 (iPhone; CPU iPhone OS 17_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Mobile/15E148 Safari/604.1",
            "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",
            "Mozilla/5.0 (iPhone; CPU iPhone OS 15_7 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.6 Mobile/15E148 Safari/604.1",

            # iPhone Chrome
            "Mozilla/5.0 (iPhone; CPU iPhone OS 17_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/120.0.6099.119 Mobile/15E148 Safari/604.1",
            "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/119.0.6045.169 Mobile/15E148 Safari/604.1",

            # iPad Safari
            "Mozilla/5.0 (iPad; CPU OS 17_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Mobile/15E148 Safari/604.1",
            "Mozilla/5.0 (iPad; CPU OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",

            # Macintosh Safari
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Safari/605.1.15",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.2 Safari/605.1.15",

            # Macintosh Chrome
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",

            # Macintosh Firefox
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101 Firefox/120.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101 Firefox/119.0",

            # M1/M2 Mac
            "Mozilla/5.0 (Macintosh; Apple M1 Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
            "Mozilla/5.0 (Macintosh; Apple M2 Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        ]

    @property
    def random(self):
        """랜덤 User-Agent 반환"""
        return random.choice(self.user_agents)

    def get_mobile_ua(self):
        """모바일 전용 User-Agent 반환"""
        mobile_uas = [ua for ua in self.user_agents if 'Mobile' in ua or 'iPhone' in ua or 'Android' in ua]
        return random.choice(mobile_uas)

    def get_desktop_ua(self):
        """데스크톱 전용 User-Agent 반환 (Mac만)"""
        desktop_uas = [ua for ua in self.user_agents if 'Macintosh' in ua and 'Mobile' not in ua]
        return random.choice(desktop_uas)


class ProxyRotator:
    def __init__(self, proxy_list=None):
        """
        프록시 로테이터 초기화
        proxy_list: ['ip:port:username:password', ...] 형태의 프록시 리스트
        """
        self.proxy_list = proxy_list if proxy_list else []
        self.proxy_cycle = itertools.cycle(self.proxy_list) if self.proxy_list else None
        self.current_proxy = None
        self.failed_proxies = set()
        self.proxy_failure_count = {}  # 프록시별 실패 횟수 추적
        self.max_failures_per_proxy = 3  # 프록시당 최대 실패 허용 횟수

    def get_next_proxy(self):
        """다음 프록시를 반환"""
        if not self.proxy_cycle:
            return None

        # 사용 가능한 프록시를 찾을 때까지 순환
        attempts = 0
        max_attempts = len(self.proxy_list) * 2  # 무한 루프 방지

        while attempts < max_attempts:
            proxy = next(self.proxy_cycle)

            # 완전히 실패한 프록시가 아니라면 사용
            if proxy not in self.failed_proxies:
                self.current_proxy = proxy
                proxy_ip = proxy.split(':')[0]
                failure_count = self.proxy_failure_count.get(proxy, 0)
                print(f"[PROXY] 현재 사용 중인 프록시: {proxy_ip} (실패 횟수: {failure_count})")
                return proxy

            attempts += 1

        # 모든 프록시가 완전히 실패했다면 실패 목록을 초기화
        if len(self.failed_proxies) == len(self.proxy_list):
            print("[WARNING] 모든 프록시가 실패했습니다. 실패 목록과 카운터를 초기화합니다.")
            self.failed_proxies.clear()
            self.proxy_failure_count.clear()

            # 첫 번째 프록시 반환
            if self.proxy_list:
                self.current_proxy = self.proxy_list[0]
                proxy_ip = self.current_proxy.split(':')[0]
                print(f"[PROXY] 초기화 후 사용 중인 프록시: {proxy_ip}")
                return self.current_proxy

        return None

    def mark_proxy_failed(self, proxy):
        """프록시를 실패로 표시 (누적 실패 관리)"""
        if proxy not in self.proxy_failure_count:
            self.proxy_failure_count[proxy] = 0

        self.proxy_failure_count[proxy] += 1
        proxy_ip = proxy.split(':')[0]

        # 최대 실패 횟수에 도달하면 완전히 제거
        if self.proxy_failure_count[proxy] >= self.max_failures_per_proxy:
            self.failed_proxies.add(proxy)
            print(f"[WARNING] 프록시 완전 실패로 제거: {proxy_ip} ({self.proxy_failure_count[proxy]}회 실패)")
        else:
            print(
                f"[WARNING] 프록시 일시 실패: {proxy_ip} ({self.proxy_failure_count[proxy]}/{self.max_failures_per_proxy} 실패)")

    def get_available_proxy_count(self):
        """사용 가능한 프록시 개수 반환"""
        if not self.proxy_list:
            return 0
        return len(self.proxy_list) - len(self.failed_proxies)

    def get_proxy_dict(self, proxy_string):
        """프록시 문자열을 requests용 딕셔너리로 변환"""
        if not proxy_string:
            return None

        parts = proxy_string.split(':')
        if len(parts) == 2:  # ip:port
            ip, port = parts
            return {
                'http': f'http://{ip}:{port}',
                'https': f'http://{ip}:{port}'
            }
        elif len(parts) == 4:  # ip:port:username:password
            ip, port, username, password = parts
            return {
                'http': f'http://{username}:{password}@{ip}:{port}',
                'https': f'http://{username}:{password}@{ip}:{port}'
            }
        return None


class ChromeDriver:
    def __init__(self, proxy_rotator=None) -> None:
        self.proxy_rotator = proxy_rotator
        self.ua = MobileUserAgent()  # 커스텀 모바일 User-Agent 사용
        self.set_options()
        self.set_driver()

    def set_options(self) -> None:
        self.options = Options()
        self.options.add_argument("--headless")
        self.options.add_argument("lang=ko_KR")

        # 모바일/Mac 전용 User-Agent 사용
        user_agent = self.ua.random
        self.options.add_argument(f"user-agent={user_agent}")
        print(f"[DEBUG] 사용 중인 User-Agent: {user_agent}")

        # 더 많은 브라우저 옵션 추가로 탐지 방지
        self.options.add_argument("--log-level=3")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-blink-features=AutomationControlled")
        self.options.add_argument("--exclude-switches=enable-automation")
        self.options.add_argument("--disable-extensions")
        self.options.add_argument("--no-first-run")
        self.options.add_argument("--disable-default-apps")
        self.options.add_argument("--disable-infobars")
        self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option('useAutomationExtension', False)

        # 프록시 설정
        if self.proxy_rotator:
            proxy = self.proxy_rotator.get_next_proxy()
            if proxy:
                parts = proxy.split(':')
                if len(parts) >= 2:
                    ip, port = parts[0], parts[1]
                    self.options.add_argument(f'--proxy-server=http://{ip}:{port}')
                    print(f"[DEBUG] Selenium 프록시 설정: {ip}:{port}")

    def set_driver(self) -> None:
        self.driver = webdriver.Chrome(options=self.options)
        # WebDriver 탐지 방지
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    def refresh_with_new_proxy(self):
        """새로운 프록시로 드라이버 재시작"""
        if hasattr(self, 'driver') and self.driver:
            self.driver.quit()
        self.set_options()
        self.set_driver()


class Coupang:
    @staticmethod
    def get_product_code(url: str) -> str:
        prod_code: str = url.split("products/")[-1].split("?")[0]
        return prod_code

    @staticmethod
    def get_soup_object(resp: rq.Response) -> bs:
        return bs(resp.text, "html.parser")

    def __del__(self) -> None:
        if hasattr(self, 'ch') and self.ch.driver:
            self.ch.driver.quit()

    def __init__(self, proxy_list=None) -> None:
        self.base_review_url: str = "https://www.coupang.com/vp/product/reviews"
        self.sd = SaveData()
        self.retries = 8  # 재시도 횟수 줄임
        self.delay_min = 2.0  # 최소 딜레이 증가
        self.delay_max = 8.0  # 최대 딜레이 증가
        self.page_delay_min = 3.0  # 페이지 간 최소 딜레이 증가
        self.page_delay_max = 10.0  # 페이지 간 최대 딜레이 증가

        # 타임아웃 관련 설정
        self.consecutive_timeouts = 0
        self.max_consecutive_timeouts = 3  # 연속 타임아웃 허용 횟수 감소
        self.long_wait_min = 300  # 긴 대기 시간 줄임 (5분)
        self.long_wait_max = 420  # 긴 대기 시간 줄임 (7분)

        # 프록시 로테이터 초기화
        self.proxy_rotator = ProxyRotator(proxy_list)

        # 모바일/Mac 전용 User-Agent 초기화
        self.ua = MobileUserAgent()

        # 더 정교한 헤더 설정
        self.base_headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "cache-control": "max-age=0",
            "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"macOS"',  # Windows 대신 macOS 사용
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "dnt": "1",
        }

        # 쿠키 저장용 세션
        self.session = rq.Session()

        # 헤더에 랜덤 User-Agent 적용
        self.update_headers()

        self.ch = ChromeDriver(self.proxy_rotator)
        self.page_title = None

    def get_realistic_headers(self):
        """실제 브라우저와 유사한 헤더 생성 (모바일/Mac 전용)"""
        headers = self.base_headers.copy()
        user_agent = self.ua.random
        headers["user-agent"] = user_agent

        # User-Agent에 따라 플랫폼 정보 조정
        if 'iPhone' in user_agent or 'iPad' in user_agent:
            headers["sec-ch-ua-platform"] = '"iOS"'
            headers["sec-ch-ua-mobile"] = "?1" if 'iPhone' in user_agent else "?0"
        elif 'Android' in user_agent:
            headers["sec-ch-ua-platform"] = '"Android"'
            headers["sec-ch-ua-mobile"] = "?1"
        elif 'Macintosh' in user_agent:
            headers["sec-ch-ua-platform"] = '"macOS"'
            headers["sec-ch-ua-mobile"] = "?0"

        # 랜덤 요소 추가
        if random.choice([True, False]):
            headers["x-requested-with"] = "XMLHttpRequest"

        # 쿠팡 특화 헤더
        headers.update({
            "x-coupang-target-market": "KR",
            "x-coupang-accept-language": "ko-KR",
        })

        return headers

    def update_headers(self):
        """헤더를 새로운 User-Agent로 업데이트"""
        self.headers = self.get_realistic_headers()
        print(f"[DEBUG] 헤더 User-Agent 업데이트: {self.headers['user-agent'][:70]}...")

    def get_session_with_proxy(self):
        """프록시가 적용된 requests 세션 반환"""
        session = rq.Session()
        session.headers.update(self.headers)

        # 더 현실적인 타임아웃 설정
        session.timeout = (10, 30)  # 연결 타임아웃 10초, 읽기 타임아웃 30초

        if self.proxy_rotator and self.proxy_rotator.proxy_list:
            proxy = self.proxy_rotator.get_next_proxy()
            if proxy:
                proxy_dict = self.proxy_rotator.get_proxy_dict(proxy)
                if proxy_dict:
                    session.proxies.update(proxy_dict)
                    print(f"[DEBUG] 요청에 프록시 적용: {proxy}")

        return session

    def warm_up_session(self, prod_code):
        """세션을 예열하여 쿠팡 사이트와의 연결을 설정"""
        try:
            print("[INFO] 세션 예열 중...")

            # 메인 페이지 먼저 방문
            main_url = "https://www.coupang.com"
            session = self.get_session_with_proxy()

            # 메인 페이지 방문
            resp = session.get(main_url, timeout=15)
            if resp.status_code == 200:
                print("[DEBUG] 메인 페이지 방문 성공")

                # 쿠키 업데이트
                self.session.cookies.update(resp.cookies)

                # 잠시 대기
                time.sleep(random.uniform(2, 4))

                # 상품 페이지 방문
                product_url = f"https://www.coupang.com/vp/products/{prod_code}"
                resp2 = session.get(product_url, timeout=15)

                if resp2.status_code == 200:
                    print("[DEBUG] 상품 페이지 방문 성공")
                    self.session.cookies.update(resp2.cookies)
                    return True

        except Exception as e:
            print(f"[WARNING] 세션 예열 실패: {e}")

        return False

    def get_product_title(self, prod_code: str) -> str:
        """상품명만 간단하게 추출"""
        url = f"https://www.coupang.com/vp/products/{prod_code}"
        print(f"[DEBUG] 상품 페이지 접속 중: {url}")

        try:
            # 여러 번 시도
            for attempt in range(3):
                try:
                    self.ch.driver.get(url=url)

                    print("[DEBUG] 페이지 로딩 대기 중...")
                    WebDriverWait(self.ch.driver, 30).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )

                    loading_delay = random.uniform(3.0, 6.0)  # 대기 시간 증가
                    print(f"[DEBUG] {loading_delay:.1f}초 대기 중...")
                    time.sleep(loading_delay)

                    page_source: str = self.ch.driver.page_source
                    soup = bs(page_source, "html.parser")

                    # 더 많은 상품명 선택자 시도
                    title_selectors = [
                        "h1.prod-buy-header__title",
                        ".prod-buy-header__title",
                        "h1[class*='title']",
                        ".product-title",
                        "h1",
                        ".prod-title",
                        "[data-testid='product-title']",
                        ".product-name"
                    ]

                    for selector in title_selectors:
                        title_elem = soup.select_one(selector)
                        if title_elem and title_elem.text.strip():
                            title = title_elem.text.strip()
                            print(f"[DEBUG] 상품명 발견: {title}")
                            return title

                    print(f"[WARNING] 시도 {attempt + 1}/3 - 상품명을 찾을 수 없습니다.")
                    if attempt < 2:  # 마지막 시도가 아니면
                        time.sleep(random.uniform(3, 5))

                except Exception as e:
                    print(f"[ERROR] 시도 {attempt + 1}/3 실패: {e}")
                    if attempt < 2:
                        time.sleep(random.uniform(3, 5))

            print("[WARNING] 모든 시도 후에도 상품명을 찾을 수 없습니다.")
            return "상품명 추출 실패"

        except Exception as e:
            print(f"[ERROR] get_product_title 에러: {e}")
            return "상품명 추출 실패"

    def is_timeout_error(self, exception) -> bool:
        """타임아웃 관련 예외인지 확인"""
        return isinstance(exception, (Timeout, ConnectTimeout, ReadTimeout)) or \
            (isinstance(exception, RequestException) and "timeout" in str(exception).lower())

    def handle_consecutive_timeouts(self) -> None:
        """연속 타임아웃 처리"""
        if self.consecutive_timeouts >= self.max_consecutive_timeouts:
            wait_time = random.uniform(self.long_wait_min, self.long_wait_max)
            wait_minutes = wait_time / 60
            print(f"[WARNING] 연속 {self.consecutive_timeouts}회 타임아웃 발생!")
            print(f"[INFO] 서버 안정화를 위해 {wait_minutes:.1f}분 대기합니다...")

            remaining_time = wait_time
            while remaining_time > 0:
                minutes_left = remaining_time / 60
                print(f"[INFO] 남은 대기 시간: {minutes_left:.1f}분")

                sleep_duration = min(30, remaining_time)
                time.sleep(sleep_duration)
                remaining_time -= sleep_duration

            print(f"[INFO] 대기 완료! 크롤링을 재개합니다.")
            self.consecutive_timeouts = 0

    def start(self) -> None:
        self.sd.create_directory()
        URL: str = self.input_review_url()

        if '#' in URL:
            URL = URL.split('#')[0]
            print(f"[DEBUG] URL fragment 제거: {URL}")

        prod_code: str = self.get_product_code(url=URL)
        print(f"[DEBUG] 상품 코드: {prod_code}")

        # 세션 예열
        self.warm_up_session(prod_code)

        try:
            self.title = self.get_product_title(prod_code=prod_code)
            print(f"[INFO] 상품명: {self.title}")
        except Exception as e:
            print(f"[ERROR] 상품명을 불러오는 도중 오류가 발생했습니다: {e}")
            self.title = "상품명 미확인"

        print(f"[INFO] 모든 리뷰 페이지를 순차적으로 크롤링합니다.")

        # 프록시 사용 중이라면 사용 가능한 프록시 수 출력
        if self.proxy_rotator and self.proxy_rotator.proxy_list:
            available_proxies = self.proxy_rotator.get_available_proxy_count()
            print(f"[INFO] 사용 가능한 프록시: {available_proxies}/{len(self.proxy_rotator.proxy_list)}개")

        success_count = 0
        current_page = 1
        consecutive_empty_pages = 0
        max_empty_pages = 5  # 연속 빈 페이지 허용 횟수 증가
        proxy_change_attempts = 0  # 프록시 교체 시도 횟수

        while consecutive_empty_pages < max_empty_pages:
            payload = {
                "productId": prod_code,
                "page": current_page,
                "size": 5,
                "sortBy": "ORDER_SCORE_ASC",
                "ratings": "",
                "q": "",
                "viRoleCode": 2,
                "ratingSummary": True,
            }

            result = self.fetch(payload=payload)

            if result:
                success_count += 1
                consecutive_empty_pages = 0
                proxy_change_attempts = 0  # 성공하면 프록시 교체 시도 횟수 리셋
            else:
                consecutive_empty_pages += 1
                print(f"[WARNING] 페이지 {current_page}에서 리뷰를 찾을 수 없습니다. ({consecutive_empty_pages}/{max_empty_pages})")

                # 연속 빈 페이지가 2개 이상이고 프록시를 사용 중이라면 프록시 상태 체크
                if (consecutive_empty_pages >= 2 and
                        self.proxy_rotator and
                        self.proxy_rotator.current_proxy and
                        proxy_change_attempts < 3):  # 최대 3번까지 프록시 교체 시도

                    available_proxies = self.proxy_rotator.get_available_proxy_count()
                    if available_proxies > 1:  # 다른 프록시가 있다면
                        print(f"[INFO] 연속 실패로 인한 프록시 교체 시도 ({proxy_change_attempts + 1}/3)")

                        # 현재 프록시를 일시 실패로 표시 (완전 제거는 아님)
                        self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                        proxy_change_attempts += 1

                        # 프록시 교체 후 현재 페이지 재시도
                        print(f"[INFO] 페이지 {current_page} 다른 프록시로 재시도...")
                        continue  # current_page 증가 없이 재시도

            current_page += 1

            if result and consecutive_empty_pages == 0:
                short_delay = random.uniform(1.0, 3.0)  # 짧은 딜레이도 증가
                time.sleep(short_delay)

            if current_page > 500:  # 최대 페이지 수 감소
                print("[INFO] 최대 페이지 수(500)에 도달했습니다.")
                break

        print(f"[INFO] 총 {success_count}개 페이지 크롤링 완료 (총 {current_page - 1}페이지 시도)")

        # 최종 프록시 상태 출력
        if self.proxy_rotator and self.proxy_rotator.proxy_list:
            available_proxies = self.proxy_rotator.get_available_proxy_count()
            failed_proxies = len(self.proxy_rotator.failed_proxies)
            print(f"[INFO] 최종 프록시 상태: 사용 가능 {available_proxies}개, 완전 실패 {failed_proxies}개")

    def fetch(self, payload: dict) -> bool:
        now_page: int = payload["page"]
        print(f"\n[INFO] Start crawling page {now_page} ...")
        attempt: int = 0
        proxy_attempts: int = 0  # 프록시 교체 횟수 추적
        max_proxy_attempts: int = min(10,
                                      len(self.proxy_rotator.proxy_list) if self.proxy_rotator else 0)  # 최대 프록시 시도 횟수

        while attempt < self.retries:
            try:
                # 매 요청마다 새로운 User-Agent 사용
                if attempt > 0:  # 재시도 시에만 User-Agent 변경
                    self.update_headers()

                session = self.get_session_with_proxy()

                # 세션에 기존 쿠키 적용
                session.cookies.update(self.session.cookies)

                # Referer 헤더 추가
                session.headers.update({
                    "Referer": f"https://www.coupang.com/vp/products/{payload['productId']}"
                })

                resp = session.get(
                    url=self.base_review_url,
                    params=payload,
                    timeout=(15, 30),  # 타임아웃 증가
                )

                self.consecutive_timeouts = 0

                if resp.status_code == 403:
                    print(f"[ERROR] HTTP 403 응답 - 프록시가 차단됨")
                    if self.proxy_rotator and self.proxy_rotator.current_proxy:
                        self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                    attempt += 1
                    continue
                elif resp.status_code != 200:
                    print(f"[ERROR] HTTP {resp.status_code} 응답")
                    attempt += 1
                    continue

                html = resp.text
                soup = bs(html, "html.parser")

                # 디버깅: 응답 내용 확인
                if now_page == 1 and len(html) < 10000:
                    print(f"[DEBUG] 응답 HTML 길이: {len(html)}")
                    print(f"[DEBUG] 응답 내용 미리보기: {html[:500]}")

                if self.page_title is None:
                    first_review = soup.select_one("article.sdp-review__article__list")
                    if first_review:
                        title_elem = first_review.select_one("div.sdp-review__article__list__info__product-info__name")
                        self.page_title = title_elem.text.strip() if title_elem else self.title
                    else:
                        self.page_title = self.title

                articles = soup.select("article.sdp-review__article__list")
                article_length = len(articles)

                if article_length == 0:
                    print(f"[WARNING] 페이지 {now_page}에서 리뷰를 찾을 수 없습니다.")

                    # 프록시 사용 중이라면 다른 프록시로 재시도
                    if self.proxy_rotator and self.proxy_rotator.current_proxy and proxy_attempts < max_proxy_attempts:
                        print(f"[INFO] 프록시 차단 가능성으로 다른 프록시로 재시도 ({proxy_attempts + 1}/{max_proxy_attempts})")

                        # 현재 프록시를 실패로 표시
                        self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                        proxy_attempts += 1
                        attempt += 1

                        # 짧은 딜레이 후 재시도
                        retry_delay = random.uniform(1.0, 3.0)
                        print(f"[DEBUG] {retry_delay:.1f}초 후 다른 프록시로 재시도...")
                        time.sleep(retry_delay)
                        continue

                    # 더 이상 시도할 프록시가 없거나 프록시를 사용하지 않는 경우
                    if now_page == 1:
                        print("[DEBUG] 첫 페이지 HTML 구조 확인:")
                        print(f"  - 전체 길이: {len(html)} 문자")
                        print(f"  - 'review' 포함 횟수: {html.lower().count('review')}")
                        print(f"  - 'article' 포함 횟수: {html.lower().count('article')}")

                        # 차단 여부 확인
                        blocked_indicators = [
                            "access denied", "blocked", "forbidden",
                            "captcha", "robot", "bot", "security", "verification"
                        ]

                        html_lower = html.lower()
                        is_blocked = False
                        for indicator in blocked_indicators:
                            if indicator in html_lower:
                                print(f"[WARNING] 차단 감지: '{indicator}' 발견")
                                is_blocked = True
                                break

                        # 차단이 감지되면 추가 재시도
                        if is_blocked and attempt < self.retries - 2:
                            print("[INFO] 차단 감지로 인한 추가 재시도...")
                            attempt += 1
                            long_delay = random.uniform(5.0, 10.0)
                            print(f"[DEBUG] {long_delay:.1f}초 대기 후 재시도...")
                            time.sleep(long_delay)
                            continue

                        alt_selectors = [
                            "article[class*='review']",
                            "div[class*='review-item']",
                            "div[class*='review-list']",
                            ".review-item",
                            "[data-review-id]"
                        ]

                        for selector in alt_selectors:
                            elements = soup.select(selector)
                            if elements:
                                print(f"[DEBUG] 대안 선택자 '{selector}' 발견: {len(elements)}개")

                    return False

                print(f"[SUCCESS] 페이지 {now_page}에서 {article_length}개 리뷰 발견")

                # 리뷰 데이터 처리 (기존 코드와 동일)
                for idx in range(article_length):
                    dict_data: dict[str, str | int] = dict()

                    review_date_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__reg-date"
                    )
                    review_date = review_date_elem.text.strip() if review_date_elem else "-"

                    user_name_elem = articles[idx].select_one(
                        "span.sdp-review__article__list__info__user__name"
                    )
                    user_name = user_name_elem.text.strip() if user_name_elem else "-"

                    rating_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__star-orange"
                    )
                    if rating_elem and rating_elem.get("data-rating"):
                        try:
                            rating = int(rating_elem.get("data-rating"))
                        except (ValueError, TypeError):
                            rating = 0
                    else:
                        rating = 0

                    prod_name_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__name"
                    )
                    prod_name = prod_name_elem.text.strip() if prod_name_elem else "-"

                    headline_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__headline"
                    )
                    headline = headline_elem.text.strip() if headline_elem else "등록된 헤드라인이 없습니다"

                    review_content_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__review__content.js_reviewArticleContent"
                    )
                    if review_content_elem:
                        review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                    else:
                        review_content_elem = articles[idx].select_one(
                            "div.sdp-review__article__list__review > div"
                        )
                        if review_content_elem:
                            review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                        else:
                            review_content = "등록된 리뷰내용이 없습니다"

                    answer_elem = articles[idx].select_one(
                        "span.sdp-review__article__list__survey__row__answer"
                    )
                    answer = answer_elem.text.strip() if answer_elem else "맛 평가 없음"

                    helpful_count_elem = articles[idx].select_one("span.js_reviewArticleHelpfulCount")
                    helpful_count = helpful_count_elem.text.strip() if helpful_count_elem else "0"

                    seller_name_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__seller_name"
                    )
                    if seller_name_elem:
                        seller_name = seller_name_elem.text.replace("판매자: ", "").strip()
                    else:
                        seller_name = "-"

                    review_images = articles[idx].select("div.sdp-review__article__list__attachment__list img")
                    image_count = len(review_images)

                    dict_data["title"] = self.page_title
                    dict_data["prod_name"] = prod_name
                    dict_data["review_date"] = review_date
                    dict_data["user_name"] = user_name
                    dict_data["rating"] = rating
                    dict_data["headline"] = headline
                    dict_data["review_content"] = review_content
                    dict_data["answer"] = answer
                    dict_data["helpful_count"] = helpful_count
                    dict_data["seller_name"] = seller_name
                    dict_data["image_count"] = image_count

                    self.sd.save(datas=dict_data)
                    print(f"[SUCCESS] 리뷰 저장 완료: {user_name} - {rating}점")

                page_delay = random.uniform(self.page_delay_min, self.page_delay_max)
                print(f"[DEBUG] 다음 페이지까지 {page_delay:.1f}초 대기...")
                time.sleep(page_delay)
                return True

            except RequestException as e:
                attempt += 1

                # 프록시 관련 오류인지 확인하고 처리
                error_str = str(e).lower()
                is_proxy_error = any(keyword in error_str for keyword in [
                    "403", "proxy", "connection", "timeout", "refused", "unreachable"
                ])

                if is_proxy_error and self.proxy_rotator and self.proxy_rotator.current_proxy:
                    self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                    print("[INFO] 프록시 오류로 인한 다른 프록시로 재시도합니다.")

                    # 사용 가능한 프록시가 있는지 확인
                    available_proxies = self.proxy_rotator.get_available_proxy_count()
                    if available_proxies > 0:
                        print(f"[INFO] 남은 사용 가능 프록시: {available_proxies}개")
                    else:
                        print("[WARNING] 사용 가능한 프록시가 없습니다.")

                if self.is_timeout_error(e):
                    self.consecutive_timeouts += 1
                    print(f"[ERROR] 타임아웃 발생 (연속 {self.consecutive_timeouts}회): {e}")

                    if self.consecutive_timeouts >= self.max_consecutive_timeouts:
                        self.handle_consecutive_timeouts()
                else:
                    self.consecutive_timeouts = 0
                    print(f"[ERROR] 네트워크 오류: {e}")

                print(f"[ERROR] Attempt {attempt}/{self.retries} failed")
                if attempt < self.retries:
                    retry_delay = random.uniform(self.delay_min, self.delay_max)
                    print(f"[DEBUG] {retry_delay:.1f}초 후 재시도...")
                    time.sleep(retry_delay)
                else:
                    print(f"[ERROR] 최대 요청 횟수 초과! 페이지 {now_page} 크롤링 실패.")
                    return False
            except Exception as e:
                print(f"[ERROR] 예상치 못한 오류 발생: {e}")
                self.consecutive_timeouts = 0
                return False

        return False

    @staticmethod
    def clear_console() -> None:
        command: str = "clear"
        if os.name in ("nt", "dos"):
            command = "cls"
        try:
            # TERM 환경변수가 설정되지 않은 경우를 처리
            if os.environ.get('TERM') is None:
                os.environ['TERM'] = 'xterm'
            os.system(command=command)
        except:
            pass  # 모든 오류 무시

    def input_review_url(self) -> str:
        while True:
            try:
                self.clear_console()
            except:
                pass

            review_url: str = input(
                "원하시는 상품의 URL 주소를 입력해주세요\n\n"
                "Ex)\n"
                "https://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906\n\n"
                "URL: "
            )
            if not review_url.strip():
                print("[ERROR] URL 주소가 입력되지 않았습니다")
                time.sleep(2)
                continue

            if "coupang.com" not in review_url:
                print("[ERROR] 올바른 쿠팡 URL을 입력해주세요")
                time.sleep(2)
                continue

            return review_url.strip()


class SaveData:
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        self.ws = self.wb.active
        self.ws.append([
            "상품명", "구매상품명", "작성일자", "구매자명", "평점",
            "헤드라인", "리뷰내용", "맛만족도", "도움수", "판매자", "이미지수"
        ])
        self.row: int = 2
        self.dir_name: str = "Coupang-reviews"
        self.create_directory()

    def create_directory(self) -> None:
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)
            print(f"[INFO] 디렉토리 생성: {self.dir_name}")

    def save(self, datas: dict[str, str | int]) -> None:
        try:
            safe_title = re.sub(r'[<>:"/\\|?*]', '_', datas["title"])
            file_name: str = os.path.join(self.dir_name, safe_title + ".xlsx")

            self.ws[f"A{self.row}"] = datas["title"]
            self.ws[f"B{self.row}"] = datas["prod_name"]
            self.ws[f"C{self.row}"] = datas["review_date"]
            self.ws[f"D{self.row}"] = datas["user_name"]
            self.ws[f"E{self.row}"] = datas["rating"]
            self.ws[f"F{self.row}"] = datas["headline"]
            self.ws[f"G{self.row}"] = datas["review_content"]
            self.ws[f"H{self.row}"] = datas["answer"]
            self.ws[f"I{self.row}"] = datas["helpful_count"]
            self.ws[f"J{self.row}"] = datas["seller_name"]
            self.ws[f"K{self.row}"] = datas["image_count"]

            self.row += 1
            self.wb.save(filename=file_name)

        except Exception as e:
            print(f"[ERROR] 데이터 저장 중 오류 발생: {e}")

    def __del__(self) -> None:
        try:
            if hasattr(self, 'wb'):
                self.wb.close()
        except:
            pass


def test_proxy(proxy_string):
    """프록시 연결 테스트"""
    try:
        parts = proxy_string.split(':')
        if len(parts) == 4:
            ip, port, username, password = parts
            proxy_dict = {
                'http': f'http://{username}:{password}@{ip}:{port}',
                'https': f'http://{username}:{password}@{ip}:{port}'
            }
        else:
            return False

        response = rq.get('http://httpbin.org/ip', proxies=proxy_dict, timeout=10)
        if response.status_code == 200:
            return True
    except:
        pass
    return False


def get_proxy_list():
    """프록시 목록 반환"""
    # 제공받은 프록시 목록
    proxy_list = [
        "198.23.239.134:6540:bwedbcvt:3ocdjeqvcfpi",
        "207.244.217.165:6712:bwedbcvt:3ocdjeqvcfpi",
        "107.172.163.27:6543:bwedbcvt:3ocdjeqvcfpi",
        "161.123.152.115:6360:bwedbcvt:3ocdjeqvcfpi",
        "23.94.138.75:6349:bwedbcvt:3ocdjeqvcfpi",
        "216.10.27.159:6837:bwedbcvt:3ocdjeqvcfpi",
        "136.0.207.84:6661:bwedbcvt:3ocdjeqvcfpi",
        "64.64.118.149:6732:bwedbcvt:3ocdjeqvcfpi",
        "142.147.128.93:6593:bwedbcvt:3ocdjeqvcfpi",
        "154.36.110.199:6853:bwedbcvt:3ocdjeqvcfpi"
    ]

    print(f"[INFO] {len(proxy_list)}개의 프록시가 준비되었습니다.")
    print("[NOTICE] 개선된 프록시 로테이션 시스템:")
    print("  ✅ 지능적 프록시 교체 (실패 시 자동 전환)")
    print("  ✅ 누적 실패 관리 (3회 실패 시 완전 제거)")
    print("  ✅ 리뷰 없음 = 프록시 차단으로 간주하여 재시도")
    print("  ✅ 연속 빈 페이지 시 프록시 교체")
    print("  ✅ Android/iPhone/Mac 전용 User-Agent 사용 (Windows 제외)")
    print()
    print("[NOTICE] HTTP 403/차단 발생 시:")
    print("  1. 자동으로 다른 프록시로 교체")
    print("  2. 프록시 없이 실행 (n 선택)")
    print("  3. VPN 사용 권장")
    print()

    # 프록시 사용 여부 확인
    use_proxy = input("프록시를 사용하시겠습니까? (Y/n): ").lower().strip()

    if use_proxy == 'n':
        print("[INFO] 프록시 없이 실행합니다.")
        return None
    else:
        print("[INFO] 프록시를 사용하여 실행합니다.")

        # 프록시 테스트 여부 확인
        test_proxies = input("프록시 연결을 테스트하시겠습니까? (y/N): ").lower().strip()

        if test_proxies == 'y':
            print("\n[INFO] 프록시 연결 테스트 중...")
            working_proxies = []

            for i, proxy in enumerate(proxy_list, 1):
                print(f"[TEST] {i}/{len(proxy_list)} - {proxy.split(':')[0]}:{proxy.split(':')[1]} 테스트 중...", end='')
                if test_proxy(proxy):
                    print(" ✅ 성공")
                    working_proxies.append(proxy)
                else:
                    print(" ❌ 실패")

            if working_proxies:
                print(f"\n[SUCCESS] {len(working_proxies)}/{len(proxy_list)}개 프록시가 정상 작동합니다.")
                print(f"[INFO] 작동하는 프록시만 사용하여 크롤링을 시작합니다.")
                return working_proxies
            else:
                print("\n[ERROR] 작동하는 프록시가 없습니다.")
                fallback = input("프록시 없이 실행하시겠습니까? (Y/n): ").lower().strip()
                if fallback != 'n':
                    print("[INFO] 프록시 없이 실행합니다.")
                    return None
                else:
                    print("[INFO] 프로그램을 종료합니다.")
                    exit(0)
        else:
            print("[INFO] 테스트 없이 모든 프록시를 사용합니다.")
            print("[INFO] 실행 중 자동으로 작동하지 않는 프록시를 제외합니다.")
            return proxy_list


if __name__ == "__main__":
    try:
        print("=" * 70)
        print("🛒 쿠팡 리뷰 크롤러 v3.5 (모바일/Mac 전용 User-Agent)")
        print("=" * 70)

        # 프록시 목록 가져오기
        proxy_list = get_proxy_list()

        coupang = Coupang(proxy_list=proxy_list)
        coupang.start()

        print("\n" + "=" * 70)
        print("✅ 크롤링이 완료되었습니다!")
        print("📁 결과 파일은 'Coupang-reviews' 폴더에서 확인하세요.")
        print("=" * 70)

    except KeyboardInterrupt:
        print("\n[INFO] 사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"\n[ERROR] 프로그램 실행 중 오류 발생: {e}")
        import traceback

        traceback.print_exc()
    finally:
        print("\n프로그램을 종료합니다.")