from bs4 import BeautifulSoup as bs
from pathlib import Path
from openpyxl import Workbook
from fake_useragent import UserAgent
from requests.exceptions import RequestException
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import os
import re
import requests as rq
import math
import sys
import random


class ChromeDriver:
    def __init__(self) -> None:
        self.set_options()
        self.set_driver()

    def set_options(self) -> None:
        self.options = Options()
        self.options.add_argument("--headless")
        self.options.add_argument("lang=ko_KR")
        self.options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
        self.options.add_argument("--log-level=3")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option("excludeSwitches", ["enable-logging"])

    def set_driver(self) -> None:
        self.driver = webdriver.Chrome(options=self.options)


class Coupang:
    @staticmethod
    def get_product_code(url: str) -> str:
        prod_code: str = url.split("products/")[-1].split("?")[0]
        return prod_code

    @staticmethod
    def get_soup_object(resp: rq.Response) -> bs:
        return bs(resp.text, "html.parser")

    def __del__(self) -> None:
        if hasattr(self, 'ch') and self.ch.driver:
            self.ch.driver.quit()

    def __init__(self) -> None:
        self.base_review_url: str = "https://www.coupang.com/vp/product/reviews"
        self.sd = SaveData()
        self.retries = 10
        self.delay_min = 0.5  # 최소 딜레이
        self.delay_max = 2.0  # 최대 딜레이
        self.page_delay_min = 1.0  # 페이지 간 최소 딜레이
        self.page_delay_max = 3.0  # 페이지 간 최대 딜레이
        self.headers = {
            "accept": "*/*",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "ko,en;q=0.9,en-US;q=0.8",
            "cookie": "_fbp=fb.1.1709172148924.2042270649; gd1=Y; delivery_toggle=false; srp_delivery_toggle=true; MARKETID=17272706554699560993959; x-coupang-accept-language=ko-KR;",
            "priority": "u=1, i",
            "sec-ch-ua": '"Chromium";v="131", "Not_A Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        }
        self.ch = ChromeDriver()
        self.page_title = None  # 상품명 캐싱용

    def get_product_title(self, prod_code: str) -> str:
        """상품명만 간단하게 추출"""
        url = f"https://www.coupang.com/vp/products/{prod_code}"
        print(f"[DEBUG] 상품 페이지 접속 중: {url}")

        try:
            self.ch.driver.get(url=url)

            # 페이지 로딩 대기
            print("[DEBUG] 페이지 로딩 대기 중...")
            WebDriverWait(self.ch.driver, 30).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            # 추가 대기 시간 (랜덤)
            loading_delay = random.uniform(2.0, 5.0)
            print(f"[DEBUG] {loading_delay:.1f}초 대기 중...")
            time.sleep(loading_delay)

            page_source: str = self.ch.driver.page_source
            soup = bs(page_source, "html.parser")

            # 상품명 추출 - 여러 선택자 시도
            title_selectors = [
                "h1.prod-buy-header__title",
                ".prod-buy-header__title",
                "h1[class*='title']",
                ".product-title",
                "h1"
            ]

            for selector in title_selectors:
                title_elem = soup.select_one(selector)
                if title_elem and title_elem.text.strip():
                    title = title_elem.text.strip()
                    print(f"[DEBUG] 상품명 발견: {title}")
                    return title

            print("[WARNING] 상품명을 찾을 수 없습니다.")
            return "상품명 추출 실패"

        except Exception as e:
            print(f"[ERROR] get_product_title 에러: {e}")
            return "상품명 추출 실패"

    def start(self) -> None:
        self.sd.create_directory()
        URL: str = self.input_review_url()

        # URL에서 fragment 제거 (#sdpReview 등)
        if '#' in URL:
            URL = URL.split('#')[0]
            print(f"[DEBUG] URL fragment 제거: {URL}")

        self.headers["Referer"] = URL
        prod_code: str = self.get_product_code(url=URL)
        print(f"[DEBUG] 상품 코드: {prod_code}")

        # 상품명만 추출
        try:
            self.title = self.get_product_title(prod_code=prod_code)
            print(f"[INFO] 상품명: {self.title}")
        except Exception as e:
            print(f"[ERROR] 상품명을 불러오는 도중 오류가 발생했습니다: {e}")
            self.title = "상품명 미확인"

        print(f"[INFO] 모든 리뷰 페이지를 순차적으로 크롤링합니다.")

        # 동적으로 모든 페이지 크롤링
        success_count = 0
        current_page = 1
        consecutive_empty_pages = 0
        max_empty_pages = 3  # 연속으로 빈 페이지 3개가 나오면 종료

        while consecutive_empty_pages < max_empty_pages:
            payload = {
                "productId": prod_code,
                "page": current_page,
                "size": 5,
                "sortBy": "ORDER_SCORE_ASC",
                "ratings": "",
                "q": "",
                "viRoleCode": 2,
                "ratingSummary": True,
            }

            result = self.fetch(payload=payload)

            if result:
                success_count += 1
                consecutive_empty_pages = 0  # 성공하면 빈 페이지 카운터 초기화
            else:
                consecutive_empty_pages += 1
                print(f"[WARNING] 페이지 {current_page}에서 리뷰를 찾을 수 없습니다. ({consecutive_empty_pages}/{max_empty_pages})")

            current_page += 1

            # 페이지 요청 간 짧은 랜덤 딜레이 (빈 페이지가 아닌 경우만)
            if result and consecutive_empty_pages == 0:
                short_delay = random.uniform(0.3, 1.0)
                time.sleep(short_delay)

            # 안전장치: 너무 많은 페이지를 크롤링하지 않도록 제한
            if current_page > 1000:
                print("[INFO] 최대 페이지 수(1000)에 도달했습니다.")
                break

        print(f"[INFO] 총 {success_count}개 페이지 크롤링 완료 (총 {current_page - 1}페이지 시도)")

    def fetch(self, payload: dict) -> bool:
        now_page: int = payload["page"]
        print(f"\n[INFO] Start crawling page {now_page} ...")
        attempt: int = 0

        while attempt < self.retries:
            try:
                resp = rq.get(
                    url=self.base_review_url,
                    headers=self.headers,
                    params=payload,
                    timeout=10,
                )

                if resp.status_code != 200:
                    print(f"[ERROR] HTTP {resp.status_code} 응답")
                    attempt += 1
                    continue

                html = resp.text
                soup = bs(html, "html.parser")

                # 상품명은 한 번만 가져오기 (첫 번째 리뷰에서)
                if self.page_title is None:
                    first_review = soup.select_one("article.sdp-review__article__list")
                    if first_review:
                        title_elem = first_review.select_one("div.sdp-review__article__list__info__product-info__name")
                        self.page_title = title_elem.text.strip() if title_elem else self.title
                    else:
                        self.page_title = self.title

                # Article Boxes
                articles = soup.select("article.sdp-review__article__list")
                article_length = len(articles)

                if article_length == 0:
                    print(f"[WARNING] 페이지 {now_page}에서 리뷰를 찾을 수 없습니다.")
                    # 첫 페이지에서 리뷰가 없으면 구조 확인
                    if now_page == 1:
                        print("[DEBUG] 첫 페이지 HTML 구조 확인:")
                        print(f"  - 전체 길이: {len(html)} 문자")
                        print(f"  - 'review' 포함 횟수: {html.lower().count('review')}")
                        print(f"  - 'article' 포함 횟수: {html.lower().count('article')}")

                        # 다른 가능한 리뷰 선택자들 확인
                        alt_selectors = [
                            "article[class*='review']",
                            "div[class*='review-item']",
                            "div[class*='review-list']",
                            ".review-item",
                            "[data-review-id]"
                        ]

                        for selector in alt_selectors:
                            elements = soup.select(selector)
                            if elements:
                                print(f"[DEBUG] 대안 선택자 '{selector}' 발견: {len(elements)}개")
                    return False

                print(f"[SUCCESS] 페이지 {now_page}에서 {article_length}개 리뷰 발견")

                for idx in range(article_length):
                    dict_data: dict[str, str | int] = dict()

                    # 리뷰 날짜
                    review_date_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__reg-date"
                    )
                    review_date = review_date_elem.text.strip() if review_date_elem else "-"

                    # 구매자 이름
                    user_name_elem = articles[idx].select_one(
                        "span.sdp-review__article__list__info__user__name"
                    )
                    user_name = user_name_elem.text.strip() if user_name_elem else "-"

                    # 평점 (data-rating 속성에서 가져오기)
                    rating_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__star-orange"
                    )
                    if rating_elem and rating_elem.get("data-rating"):
                        try:
                            rating = int(rating_elem.get("data-rating"))
                        except (ValueError, TypeError):
                            rating = 0
                    else:
                        rating = 0

                    # 구매자 상품명
                    prod_name_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__name"
                    )
                    prod_name = prod_name_elem.text.strip() if prod_name_elem else "-"

                    # 헤드라인(타이틀)
                    headline_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__headline"
                    )
                    headline = headline_elem.text.strip() if headline_elem else "등록된 헤드라인이 없습니다"

                    # 리뷰 내용 - 개선된 선택자 사용
                    review_content_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__review__content.js_reviewArticleContent"
                    )
                    if review_content_elem:
                        review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                    else:
                        # 백업 선택자 시도
                        review_content_elem = articles[idx].select_one(
                            "div.sdp-review__article__list__review > div"
                        )
                        if review_content_elem:
                            review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                        else:
                            review_content = "등록된 리뷰내용이 없습니다"

                    # 맛 만족도
                    answer_elem = articles[idx].select_one(
                        "span.sdp-review__article__list__survey__row__answer"
                    )
                    answer = answer_elem.text.strip() if answer_elem else "맛 평가 없음"

                    # 추가 정보들
                    # 도움이 된 사람 수
                    helpful_count_elem = articles[idx].select_one("span.js_reviewArticleHelpfulCount")
                    helpful_count = helpful_count_elem.text.strip() if helpful_count_elem else "0"

                    # 판매자 정보
                    seller_name_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__seller_name"
                    )
                    if seller_name_elem:
                        seller_name = seller_name_elem.text.replace("판매자: ", "").strip()
                    else:
                        seller_name = "-"

                    # 리뷰 이미지 개수
                    review_images = articles[idx].select("div.sdp-review__article__list__attachment__list img")
                    image_count = len(review_images)

                    # 데이터 저장
                    dict_data["title"] = self.page_title
                    dict_data["prod_name"] = prod_name
                    dict_data["review_date"] = review_date
                    dict_data["user_name"] = user_name
                    dict_data["rating"] = rating
                    dict_data["headline"] = headline
                    dict_data["review_content"] = review_content
                    dict_data["answer"] = answer
                    dict_data["helpful_count"] = helpful_count
                    dict_data["seller_name"] = seller_name
                    dict_data["image_count"] = image_count

                    self.sd.save(datas=dict_data)
                    print(f"[SUCCESS] 리뷰 저장 완료: {user_name} - {rating}점")

                # 페이지 간 랜덤 딜레이
                page_delay = random.uniform(self.page_delay_min, self.page_delay_max)
                print(f"[DEBUG] 다음 페이지까지 {page_delay:.1f}초 대기...")
                time.sleep(page_delay)
                return True

            except RequestException as e:
                attempt += 1
                print(f"[ERROR] Attempt {attempt}/{self.retries} failed: {e}")
                if attempt < self.retries:
                    retry_delay = random.uniform(self.delay_min, self.delay_max)
                    print(f"[DEBUG] {retry_delay:.1f}초 후 재시도...")
                    time.sleep(retry_delay)
                else:
                    print(f"[ERROR] 최대 요청 횟수 초과! 페이지 {now_page} 크롤링 실패.")
                    return False
            except Exception as e:
                print(f"[ERROR] 예상치 못한 오류 발생: {e}")
                return False

        return False

    @staticmethod
    def clear_console() -> None:
        command: str = "clear"
        if os.name in ("nt", "dos"):
            command = "cls"
        try:
            os.system(command=command)
        except:
            pass  # TERM 환경변수 오류 무시

    def input_review_url(self) -> str:
        while True:
            try:
                self.clear_console()
            except:
                pass

            review_url: str = input(
                "원하시는 상품의 URL 주소를 입력해주세요\n\n"
                "Ex)\n"
                "https://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906\n\n"
                "URL: "
            )
            if not review_url.strip():
                print("[ERROR] URL 주소가 입력되지 않았습니다")
                time.sleep(2)
                continue

            if "coupang.com" not in review_url:
                print("[ERROR] 올바른 쿠팡 URL을 입력해주세요")
                time.sleep(2)
                continue

            return review_url.strip()


class SaveData:
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        self.ws = self.wb.active
        # 헤더 업데이트 - 새로운 필드들 추가
        self.ws.append([
            "상품명", "구매상품명", "작성일자", "구매자명", "평점",
            "헤드라인", "리뷰내용", "맛만족도", "도움수", "판매자", "이미지수"
        ])
        self.row: int = 2
        self.dir_name: str = "Coupang-reviews"
        self.create_directory()

    def create_directory(self) -> None:
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)
            print(f"[INFO] 디렉토리 생성: {self.dir_name}")

    def save(self, datas: dict[str, str | int]) -> None:
        try:
            # 파일명에 사용할 수 없는 문자 제거
            safe_title = re.sub(r'[<>:"/\\|?*]', '_', datas["title"])
            file_name: str = os.path.join(self.dir_name, safe_title + ".xlsx")

            self.ws[f"A{self.row}"] = datas["title"]
            self.ws[f"B{self.row}"] = datas["prod_name"]
            self.ws[f"C{self.row}"] = datas["review_date"]
            self.ws[f"D{self.row}"] = datas["user_name"]
            self.ws[f"E{self.row}"] = datas["rating"]
            self.ws[f"F{self.row}"] = datas["headline"]
            self.ws[f"G{self.row}"] = datas["review_content"]
            self.ws[f"H{self.row}"] = datas["answer"]
            self.ws[f"I{self.row}"] = datas["helpful_count"]
            self.ws[f"J{self.row}"] = datas["seller_name"]
            self.ws[f"K{self.row}"] = datas["image_count"]

            self.row += 1
            self.wb.save(filename=file_name)

        except Exception as e:
            print(f"[ERROR] 데이터 저장 중 오류 발생: {e}")

    def __del__(self) -> None:
        try:
            if hasattr(self, 'wb'):
                self.wb.close()
        except:
            pass


if __name__ == "__main__":
    try:
        print("=" * 50)
        print("🛒 쿠팡 리뷰 크롤러 v3.1 (랜덤 딜레이 적용)")
        print("=" * 50)

        coupang = Coupang()
        coupang.start()

        print("\n" + "=" * 50)
        print("✅ 크롤링이 완료되었습니다!")
        print("📁 결과 파일은 'Coupang-reviews' 폴더에서 확인하세요.")
        print("=" * 50)

    except KeyboardInterrupt:
        print("\n[INFO] 사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"\n[ERROR] 프로그램 실행 중 오류 발생: {e}")
        import traceback

        traceback.print_exc()
    finally:
        print("\n프로그램을 종료합니다.")