import os
import re
from openpyxl import Workbook
from itemadapter import ItemAdapter


class CoupangReviewsPipeline:
    """
    기존 코드의 Excel 저장 로직을 Scrapy 파이프라인으로 구현
    """

    def __init__(self):
        self.wb = None
        self.ws = None
        self.row = 2
        self.dir_name = "Coupang-reviews"
        self.current_file = None

    def open_spider(self, spider):
        """스파이더 시작 시 호출"""
        self.create_directory()
        spider.logger.info(f"파이프라인 시작 - 저장 폴더: {self.dir_name}")

    def close_spider(self, spider):
        """스파이더 종료 시 호출"""
        if self.wb:
            try:
                self.wb.save(filename=self.current_file)
                self.wb.close()
                spider.logger.info(f"Excel 파일 저장 완료: {self.current_file}")
            except Exception as e:
                spider.logger.error(f"Excel 파일 저장 중 오류: {e}")

    def process_item(self, item, spider):
        """각 아이템 처리"""
        adapter = ItemAdapter(item)

        # 첫 번째 아이템에서 워크북 초기화
        if self.wb is None:
            self._init_workbook(adapter.get('title', '상품명_미확인'))

        try:
            # Excel에 데이터 저장
            self.ws[f"A{self.row}"] = adapter.get('title', '')
            self.ws[f"B{self.row}"] = adapter.get('prod_name', '')
            self.ws[f"C{self.row}"] = adapter.get('review_date', '')
            self.ws[f"D{self.row}"] = adapter.get('user_name', '')
            self.ws[f"E{self.row}"] = adapter.get('rating', 0)
            self.ws[f"F{self.row}"] = adapter.get('headline', '')
            self.ws[f"G{self.row}"] = adapter.get('review_content', '')
            self.ws[f"H{self.row}"] = adapter.get('answer', '')
            self.ws[f"I{self.row}"] = adapter.get('helpful_count', '')
            self.ws[f"J{self.row}"] = adapter.get('seller_name', '')
            self.ws[f"K{self.row}"] = adapter.get('image_count', 0)

            self.row += 1

            # 주기적으로 저장 (10개마다)
            if self.row % 10 == 0:
                self.wb.save(filename=self.current_file)

            spider.logger.info(f"리뷰 저장 완료: {adapter.get('user_name')} - {adapter.get('rating')}점")

        except Exception as e:
            spider.logger.error(f"데이터 저장 중 오류 발생: {e}")

        return item

    def create_directory(self):
        """디렉토리 생성"""
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)

    def _init_workbook(self, title):
        """워크북 초기화"""
        self.wb = Workbook()
        self.ws = self.wb.active

        # 헤더 추가
        self.ws.append([
            "상품명", "구매상품명", "작성일자", "구매자명", "평점",
            "헤드라인", "리뷰내용", "맛만족도", "도움수", "판매자", "이미지수"
        ])

        # 파일명 설정
        safe_title = re.sub(r'[<>:"/\\|?*]', '_', title)
        self.current_file = os.path.join(self.dir_name, safe_title + ".xlsx")