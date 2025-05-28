from bs4 import BeautifulSoup as bs
from pathlib import Path
from openpyxl import Workbook
from fake_useragent import UserAgent
from requests.exceptions import RequestException, Timeout, ConnectionError
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import os
import re
import requests as rq
import math
import sys
import random


class ChromeDriver:
    def __init__(self) -> None:
        self.set_options()
        self.set_driver()

    def set_options(self) -> None:
        self.options = Options()
        self.options.add_argument("--headless")
        self.options.add_argument("lang=ko_KR")
        self.options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        )
        self.options.add_argument("--log-level=3")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option("excludeSwitches", ["enable-logging"])

    def set_driver(self) -> None:
        self.driver = webdriver.Chrome(options=self.options)


class Coupang:
    @staticmethod
    def get_product_code(url: str) -> str:
        prod_code: str = url.split("products/")[-1].split("?")[0]
        return prod_code

    @staticmethod
    def get_soup_object(resp: rq.Response) -> bs:
        return bs(resp.text, "html.parser")

    def __del__(self) -> None:
        if hasattr(self, 'ch') and self.ch.driver:
            self.ch.driver.quit()
        if hasattr(self, 'session') and self.session:
            self.session.close()

    def __init__(self) -> None:
        self.base_review_url: str = "https://www.coupang.com/vp/product/reviews"
        self.sd = SaveData()
        self.retries = 6  # 재시도 횟수 줄임 (403 오류 시 많은 재시도는 무의미)
        self.delay_min = 1.5  # 최소 딜레이 증가
        self.delay_max = 4.0  # 최대 딜레이 증가
        self.page_delay_min = 2.0  # 페이지 간 최소 딜레이 증가
        self.page_delay_max = 5.0  # 페이지 간 최대 딜레이 증가

        # 최적화된 타임아웃 설정
        self.timeout_connect = 10
        self.timeout_read = 20

        # 더 자연스러운 헤더 구성
        self.base_headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "cache-control": "max-age=0",
            "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        }

        self.ajax_headers = {
            "accept": "*/*",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "x-requested-with": "XMLHttpRequest",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        }

        self.ch = ChromeDriver()
        self.page_title = None
        self.session = None
        self.main_url = None  # 메인 상품 URL 저장

        # 최신 User-Agent 목록
        self.user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:132.0) Gecko/20100101 Firefox/132.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15"
        ]
        self.current_ua_index = 0
        self.setup_session()

    def setup_session(self) -> None:
        """브라우저와 유사한 세션 설정"""
        self.session = rq.Session()

        # 기본 헤더 설정
        self.session.headers.update(self.base_headers)

        print("[INFO] 브라우저 세션이 설정되었습니다.")

    def get_next_user_agent(self) -> str:
        """순차적으로 User-Agent 변경"""
        ua = self.user_agents[self.current_ua_index]
        self.current_ua_index = (self.current_ua_index + 1) % len(self.user_agents)
        return ua

    def warm_up_session(self, product_url: str) -> bool:
        """실제 브라우저처럼 메인 페이지 먼저 방문하여 세션 준비"""
        try:
            print("[INFO] 세션 준비 중 - 메인 페이지 방문...")

            # 먼저 쿠팡 메인 페이지 방문
            main_resp = self.session.get(
                "https://www.coupang.com",
                timeout=(self.timeout_connect, self.timeout_read)
            )

            if main_resp.status_code != 200:
                print(f"[WARNING] 메인 페이지 방문 실패: {main_resp.status_code}")
            else:
                print("[SUCCESS] 메인 페이지 방문 완료")

            # 잠시 대기
            time.sleep(random.uniform(1.0, 2.0))

            # 상품 페이지 방문
            print("[INFO] 상품 페이지 방문 중...")
            product_resp = self.session.get(
                product_url,
                timeout=(self.timeout_connect, self.timeout_read)
            )

            if product_resp.status_code != 200:
                print(f"[WARNING] 상품 페이지 방문 실패: {product_resp.status_code}")
                return False
            else:
                print("[SUCCESS] 상품 페이지 방문 완료")

                # 세션 쿠키 및 헤더 업데이트
                if 'Set-Cookie' in product_resp.headers:
                    print("[INFO] 세션 쿠키 설정 완료")

                # Referer 설정
                self.ajax_headers["referer"] = product_url

                return True

        except Exception as e:
            print(f"[ERROR] 세션 준비 중 오류: {e}")
            return False

    def get_product_title(self, prod_code: str) -> str:
        """상품명만 간단하게 추출"""
        url = f"https://www.coupang.com/vp/products/{prod_code}"
        self.main_url = url  # 메인 URL 저장
        print(f"[DEBUG] 상품 페이지 접속 중: {url}")

        try:
            self.ch.driver.get(url=url)

            # 페이지 로딩 대기
            print("[DEBUG] 페이지 로딩 대기 중...")
            WebDriverWait(self.ch.driver, 30).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            # 추가 대기 시간 (랜덤)
            loading_delay = random.uniform(3.0, 6.0)  # 더 긴 대기 시간
            print(f"[DEBUG] {loading_delay:.1f}초 대기 중...")
            time.sleep(loading_delay)

            page_source: str = self.ch.driver.page_source
            soup = bs(page_source, "html.parser")

            # 상품명 추출 - 여러 선택자 시도
            title_selectors = [
                "h1.prod-buy-header__title",
                ".prod-buy-header__title",
                "h1[class*='title']",
                ".product-title",
                "h1"
            ]

            for selector in title_selectors:
                title_elem = soup.select_one(selector)
                if title_elem and title_elem.text.strip():
                    title = title_elem.text.strip()
                    print(f"[DEBUG] 상품명 발견: {title}")
                    return title

            print("[WARNING] 상품명을 찾을 수 없습니다.")
            return "상품명 추출 실패"

        except Exception as e:
            print(f"[ERROR] get_product_title 에러: {e}")
            return "상품명 추출 실패"

    def start(self) -> None:
        self.sd.create_directory()
        URL: str = self.input_review_url()

        # URL에서 fragment 제거 (#sdpReview 등)
        if '#' in URL:
            URL = URL.split('#')[0]
            print(f"[DEBUG] URL fragment 제거: {URL}")

        prod_code: str = self.get_product_code(url=URL)
        print(f"[DEBUG] 상품 코드: {prod_code}")

        # 상품명 추출
        try:
            self.title = self.get_product_title(prod_code=prod_code)
            print(f"[INFO] 상품명: {self.title}")
        except Exception as e:
            print(f"[ERROR] 상품명을 불러오는 도중 오류가 발생했습니다: {e}")
            self.title = "상품명 미확인"

        # 세션 준비 (중요!)
        if not self.warm_up_session(self.main_url):
            print("[ERROR] 세션 준비에 실패했습니다. 계속 진행하지만 403 오류가 발생할 수 있습니다.")

        print(f"[INFO] 모든 리뷰 페이지를 순차적으로 크롤링합니다.")

        # 동적으로 모든 페이지 크롤링
        success_count = 0
        current_page = 1
        consecutive_empty_pages = 0
        max_empty_pages = 3
        consecutive_403_count = 0  # 연속 403 오류 카운터
        max_403_count = 5  # 연속 403 오류 최대 허용 횟수

        while consecutive_empty_pages < max_empty_pages and consecutive_403_count < max_403_count:
            payload = {
                "productId": prod_code,
                "page": current_page,
                "size": 5,
                "sortBy": "ORDER_SCORE_ASC",
                "ratings": "",
                "q": "",
                "viRoleCode": 2,
                "ratingSummary": True,
            }

            result = self.fetch(payload=payload)

            if result == "403_error":
                consecutive_403_count += 1
                print(f"[WARNING] 연속 403 오류 발생 ({consecutive_403_count}/{max_403_count})")
                if consecutive_403_count >= max_403_count:
                    print("[ERROR] 연속 403 오류로 인해 크롤링을 중단합니다.")
                    break
                # 403 오류 시 더 긴 대기
                long_delay = random.uniform(10.0, 20.0)
                print(f"[INFO] 403 오류 복구를 위해 {long_delay:.1f}초 대기...")
                time.sleep(long_delay)
            elif result:
                success_count += 1
                consecutive_empty_pages = 0
                consecutive_403_count = 0  # 성공 시 403 카운터 리셋
            else:
                consecutive_empty_pages += 1
                print(f"[WARNING] 페이지 {current_page}에서 리뷰를 찾을 수 없습니다. ({consecutive_empty_pages}/{max_empty_pages})")

            current_page += 1

            # 페이지 요청 간 더 긴 랜덤 딜레이
            if result and result != "403_error":
                short_delay = random.uniform(1.0, 3.0)  # 더 긴 딜레이
                time.sleep(short_delay)

            # 안전장치
            if current_page > 1000:
                print("[INFO] 최대 페이지 수(1000)에 도달했습니다.")
                break

        print(f"[INFO] 총 {success_count}개 페이지 크롤링 완료 (총 {current_page - 1}페이지 시도)")

    def fetch(self, payload: dict) -> bool | str:
        now_page: int = payload["page"]
        print(f"\n[INFO] Start crawling page {now_page} ...")
        attempt: int = 0

        while attempt < self.retries:
            try:
                # 주기적으로 User-Agent 및 헤더 업데이트
                if now_page % 3 == 1:  # 3페이지마다
                    current_ua = self.get_next_user_agent()
                    self.ajax_headers["user-agent"] = current_ua
                    print(f"[DEBUG] User-Agent 변경: {current_ua[:50]}...")

                # AJAX 헤더로 요청
                resp = self.session.get(
                    url=self.base_review_url,
                    params=payload,
                    headers=self.ajax_headers,
                    timeout=(self.timeout_connect, self.timeout_read),
                )

                if resp.status_code == 403:
                    print(f"[ERROR] HTTP 403 응답 - 봇 차단 감지")
                    return "403_error"
                elif resp.status_code != 200:
                    print(f"[ERROR] HTTP {resp.status_code} 응답")
                    attempt += 1
                    time.sleep(random.uniform(2.0, 4.0))
                    continue

                html = resp.text
                soup = bs(html, "html.parser")

                # 상품명은 한 번만 가져오기
                if self.page_title is None:
                    first_review = soup.select_one("article.sdp-review__article__list")
                    if first_review:
                        title_elem = first_review.select_one("div.sdp-review__article__list__info__product-info__name")
                        self.page_title = title_elem.text.strip() if title_elem else self.title
                    else:
                        self.page_title = self.title

                # Article Boxes
                articles = soup.select("article.sdp-review__article__list")
                article_length = len(articles)

                if article_length == 0:
                    print(f"[WARNING] 페이지 {now_page}에서 리뷰를 찾을 수 없습니다.")
                    if now_page == 1:
                        print("[DEBUG] 첫 페이지 HTML 구조 확인:")
                        print(f"  - 전체 길이: {len(html)} 문자")
                        print(f"  - 'review' 포함 횟수: {html.lower().count('review')}")
                        print(f"  - 'article' 포함 횟수: {html.lower().count('article')}")

                        # 다른 가능한 리뷰 선택자들 확인
                        alt_selectors = [
                            "article[class*='review']",
                            "div[class*='review-item']",
                            "div[class*='review-list']",
                            ".review-item",
                            "[data-review-id]"
                        ]

                        for selector in alt_selectors:
                            elements = soup.select(selector)
                            if elements:
                                print(f"[DEBUG] 대안 선택자 '{selector}' 발견: {len(elements)}개")
                    return False

                print(f"[SUCCESS] 페이지 {now_page}에서 {article_length}개 리뷰 발견")

                for idx in range(article_length):
                    dict_data: dict[str, str | int] = dict()

                    # 리뷰 데이터 추출 (기존과 동일)
                    review_date_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__reg-date"
                    )
                    review_date = review_date_elem.text.strip() if review_date_elem else "-"

                    user_name_elem = articles[idx].select_one(
                        "span.sdp-review__article__list__info__user__name"
                    )
                    user_name = user_name_elem.text.strip() if user_name_elem else "-"

                    rating_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__star-orange"
                    )
                    if rating_elem and rating_elem.get("data-rating"):
                        try:
                            rating = int(rating_elem.get("data-rating"))
                        except (ValueError, TypeError):
                            rating = 0
                    else:
                        rating = 0

                    prod_name_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__name"
                    )
                    prod_name = prod_name_elem.text.strip() if prod_name_elem else "-"

                    headline_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__headline"
                    )
                    headline = headline_elem.text.strip() if headline_elem else "등록된 헤드라인이 없습니다"

                    review_content_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__review__content.js_reviewArticleContent"
                    )
                    if review_content_elem:
                        review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                    else:
                        review_content_elem = articles[idx].select_one(
                            "div.sdp-review__article__list__review > div"
                        )
                        if review_content_elem:
                            review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                        else:
                            review_content = "등록된 리뷰내용이 없습니다"

                    answer_elem = articles[idx].select_one(
                        "span.sdp-review__article__list__survey__row__answer"
                    )
                    answer = answer_elem.text.strip() if answer_elem else "맛 평가 없음"

                    helpful_count_elem = articles[idx].select_one("span.js_reviewArticleHelpfulCount")
                    helpful_count = helpful_count_elem.text.strip() if helpful_count_elem else "0"

                    seller_name_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__seller_name"
                    )
                    if seller_name_elem:
                        seller_name = seller_name_elem.text.replace("판매자: ", "").strip()
                    else:
                        seller_name = "-"

                    review_images = articles[idx].select("div.sdp-review__article__list__attachment__list img")
                    image_count = len(review_images)

                    # 데이터 저장
                    dict_data["title"] = self.page_title
                    dict_data["prod_name"] = prod_name
                    dict_data["review_date"] = review_date
                    dict_data["user_name"] = user_name
                    dict_data["rating"] = rating
                    dict_data["headline"] = headline
                    dict_data["review_content"] = review_content
                    dict_data["answer"] = answer
                    dict_data["helpful_count"] = helpful_count
                    dict_data["seller_name"] = seller_name
                    dict_data["image_count"] = image_count

                    self.sd.save(datas=dict_data)
                    print(f"[SUCCESS] 리뷰 저장 완료: {user_name} - {rating}점")

                # 페이지 간 더 긴 랜덤 딜레이
                page_delay = random.uniform(self.page_delay_min, self.page_delay_max)
                print(f"[DEBUG] 다음 페이지까지 {page_delay:.1f}초 대기...")
                time.sleep(page_delay)
                return True

            except Timeout as e:
                attempt += 1
                print(f"[ERROR] 타임아웃 발생 (시도 {attempt}/{self.retries}): {e}")
                if attempt < self.retries:
                    retry_delay = random.uniform(3.0, 6.0)
                    print(f"[DEBUG] 타임아웃 복구를 위해 {retry_delay:.1f}초 대기...")
                    time.sleep(retry_delay)
                else:
                    print(f"[ERROR] 타임아웃으로 인한 페이지 {now_page} 크롤링 실패.")
                    return False

            except ConnectionError as e:
                attempt += 1
                print(f"[ERROR] 연결 오류 발생 (시도 {attempt}/{self.retries}): {e}")
                if attempt < self.retries:
                    retry_delay = random.uniform(4.0, 8.0)
                    print(f"[DEBUG] 연결 복구를 위해 {retry_delay:.1f}초 대기...")
                    time.sleep(retry_delay)
                else:
                    print(f"[ERROR] 연결 오류로 인한 페이지 {now_page} 크롤링 실패.")
                    return False

            except RequestException as e:
                attempt += 1
                print(f"[ERROR] 요청 오류 발생 (시도 {attempt}/{self.retries}): {e}")
                if attempt < self.retries:
                    retry_delay = random.uniform(self.delay_min, self.delay_max)
                    print(f"[DEBUG] {retry_delay:.1f}초 후 재시도...")
                    time.sleep(retry_delay)
                else:
                    print(f"[ERROR] 최대 요청 횟수 초과! 페이지 {now_page} 크롤링 실패.")
                    return False
            except Exception as e:
                print(f"[ERROR] 예상치 못한 오류 발생: {e}")
                return False

        return False

    @staticmethod
    def clear_console() -> None:
        command: str = "clear"
        if os.name in ("nt", "dos"):
            command = "cls"
        try:
            os.system(command=command)
        except:
            pass

    def input_review_url(self) -> str:
        while True:
            try:
                self.clear_console()
            except:
                pass

            review_url: str = input(
                "원하시는 상품의 URL 주소를 입력해주세요\n\n"
                "Ex)\n"
                "https://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906\n\n"
                "URL: "
            )
            if not review_url.strip():
                print("[ERROR] URL 주소가 입력되지 않았습니다")
                time.sleep(2)
                continue

            if "coupang.com" not in review_url:
                print("[ERROR] 올바른 쿠팡 URL을 입력해주세요")
                time.sleep(2)
                continue

            return review_url.strip()


class SaveData:
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        self.ws = self.wb.active
        self.ws.append([
            "상품명", "구매상품명", "작성일자", "구매자명", "평점",
            "헤드라인", "리뷰내용", "맛만족도", "도움수", "판매자", "이미지수"
        ])
        self.row: int = 2
        self.dir_name: str = "Coupang-reviews"
        self.create_directory()

    def create_directory(self) -> None:
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)
            print(f"[INFO] 디렉토리 생성: {self.dir_name}")

    def save(self, datas: dict[str, str | int]) -> None:
        try:
            safe_title = re.sub(r'[<>:"/\\|?*]', '_', datas["title"])
            file_name: str = os.path.join(self.dir_name, safe_title + ".xlsx")

            self.ws[f"A{self.row}"] = datas["title"]
            self.ws[f"B{self.row}"] = datas["prod_name"]
            self.ws[f"C{self.row}"] = datas["review_date"]
            self.ws[f"D{self.row}"] = datas["user_name"]
            self.ws[f"E{self.row}"] = datas["rating"]
            self.ws[f"F{self.row}"] = datas["headline"]
            self.ws[f"G{self.row}"] = datas["review_content"]
            self.ws[f"H{self.row}"] = datas["answer"]
            self.ws[f"I{self.row}"] = datas["helpful_count"]
            self.ws[f"J{self.row}"] = datas["seller_name"]
            self.ws[f"K{self.row}"] = datas["image_count"]

            self.row += 1
            self.wb.save(filename=file_name)

        except Exception as e:
            print(f"[ERROR] 데이터 저장 중 오류 발생: {e}")

    def __del__(self) -> None:
        try:
            if hasattr(self, 'wb'):
                self.wb.close()
        except:
            pass


if __name__ == "__main__":
    try:
        print("=" * 50)
        print("🛒 쿠팡 리뷰 크롤러 v3.3 (HTTP 403 해결)")
        print("=" * 50)

        coupang = Coupang()
        coupang.start()

        print("\n" + "=" * 50)
        print("✅ 크롤링이 완료되었습니다!")
        print("📁 결과 파일은 'Coupang-reviews' 폴더에서 확인하세요.")
        print("=" * 50)

    except KeyboardInterrupt:
        print("\n[INFO] 사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"\n[ERROR] 프로그램 실행 중 오류 발생: {e}")
        import traceback

        traceback.print_exc()
    finally:
        print("\n프로그램을 종료합니다.")